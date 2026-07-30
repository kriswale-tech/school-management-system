import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from academics.services.curriculum import provision_school_curriculum, seed_ghana_curriculum
from accounts.models import SchoolMembership, User
from accounts.tests.factories import (
    create_school,
    create_user,
    set_client_auth_cookies,
    user_school,
)
from schools.models import AcademicYear, SchoolSetup, Term
from schools.services.teachers_bulk_import import get_failure_export_path
from schools.services.teachers_bulk_template import build_teacher_failure_xlsx
from schools.tests.factories import create_school_setup
from shared.services.spreadsheets import build_csv_bytes
from teachers.models import ClassTeacher, TeachingAssignment


class TeacherBulkImportTests(APITestCase):
    def setUp(self):
        self.admin = create_user(is_active=True)
        set_client_auth_cookies(self.client, self.admin)
        self.school = user_school(self.admin)
        seed_ghana_curriculum()
        provision_school_curriculum(self.school)
        self.academic_year = AcademicYear.objects.create(
            school=self.school,
            academic_year='2025/2026',
            start_date='2025-09-01',
            end_date='2026-07-31',
            is_active=True,
        )
        self.term = Term.objects.create(
            school=self.school,
            academic_year=self.academic_year,
            term=Term.TermChoices.FIRST_TERM,
            start_date='2025-09-01',
            end_date='2025-12-15',
            is_active=True,
        )
        create_school_setup(
            self.school,
            completed_steps=[
                SchoolSetup.SetupStep.SCHOOL_PROFILE,
                SchoolSetup.SetupStep.ACADEMIC_YEAR_TERM,
                SchoolSetup.SetupStep.CLASSES_AND_SUBJECTS,
                SchoolSetup.SetupStep.ASSESSMENT,
                SchoolSetup.SetupStep.FEES,
            ],
            current_step=SchoolSetup.SetupStep.TEACHERS,
        )
        self.class_level = self.school.class_levels.get(name='Basic 4')
        self.class_subject = self.school.class_subjects.filter(
            class_level=self.class_level,
            subject__name='Mathematics',
        ).first()
        self.template_url = reverse('school-setup-teachers-bulk-upload-template')
        self.upload_url = reverse('school-setup-teachers-bulk-upload')

    def test_download_template_returns_xlsx(self):
        response = self.client.get(self.template_url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        content = b''.join(response.streaming_content)
        workbook = load_workbook(filename=io.BytesIO(content))
        self.assertIn('Import', workbook.sheetnames)
        self.assertIn('Reference', workbook.sheetnames)

    def test_dry_run_reports_invalid_subject(self):
        upload = self._csv_upload([
            {
                'first_name': 'Ama',
                'last_name': 'Boateng',
                'phone_number': '0244567891',
                'assignment_type': 'teaching',
                'class_name': 'Basic 4',
                'subject_name': 'Maths',
            },
        ])

        response = self.client.post(f'{self.upload_url}?dry_run=true', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['dry_run'])
        self.assertEqual(response.data['summary']['rows_with_errors'], 1)
        self.assertEqual(response.data['rows'][0]['status'], 'error')

    def test_dry_run_reports_teacher_from_another_school_as_a_link(self):
        create_user(
            phone_number='+233244567891',
            email='ama@other.com',
            school=create_school(name='Other School', phone_number='+233200000000'),
            role=User.RoleChoices.TEACHER,
            is_active=True,
            first_name='Ama',
            last_name='Boateng',
        )

        upload = self._csv_upload([
            {
                'first_name': 'Ama',
                'last_name': 'Boateng',
                'phone_number': '0244567891',
                'assignment_type': 'class_teacher',
                'class_name': 'Basic 4',
            },
        ])

        response = self.client.post(
            f'{self.upload_url}?dry_run=true',
            {'file': upload},
            format='multipart',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['summary']['rows_with_errors'], 0)
        self.assertEqual(response.data['summary']['teachers_to_link'], 1)
        self.assertEqual(response.data['summary']['teachers_to_create'], 0)

    def test_commit_links_teacher_from_another_school_without_renaming_them(self):
        existing = create_user(
            phone_number='+233244567891',
            email='ama@other.com',
            school=create_school(name='Other School', phone_number='+233200000000'),
            role=User.RoleChoices.TEACHER,
            is_active=True,
            first_name='Ama',
            last_name='Boateng',
        )

        upload = self._csv_upload([
            {
                'first_name': 'Wrong',
                'last_name': 'Name',
                'phone_number': '0244567891',
                'assignment_type': 'class_teacher',
                'class_name': 'Basic 4',
            },
        ])

        response = self.client.post(
            f'{self.upload_url}?dry_run=false',
            {'file': upload},
            format='multipart',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['summary']['rows_succeeded'], 1)
        self.assertEqual(response.data['summary']['teachers_linked'], 1)
        self.assertEqual(User.objects.filter(phone_number='+233244567891').count(), 1)

        existing.refresh_from_db()
        self.assertEqual(existing.first_name, 'Ama')
        self.assertTrue(
            SchoolMembership.objects.filter(
                user=existing,
                school=self.school,
                role=User.RoleChoices.TEACHER,
                is_active=True,
            ).exists(),
        )
        self.assertEqual(
            ClassTeacher.objects.filter(teacher=existing, term=self.term).count(),
            1,
        )

    def test_commit_creates_teacher_and_assignments(self):
        upload = self._csv_upload([
            {
                'first_name': 'Ama',
                'last_name': 'Boateng',
                'phone_number': '0244567891',
                'email': 'ama@school.com',
                'assignment_type': 'class_teacher',
                'class_name': 'Basic 4',
            },
            {
                'first_name': 'Ama',
                'last_name': 'Boateng',
                'phone_number': '0244567891',
                'assignment_type': 'teaching',
                'class_name': 'Basic 4',
                'subject_name': 'Mathematics',
            },
        ])

        response = self.client.post(f'{self.upload_url}?dry_run=false', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data['dry_run'])
        self.assertEqual(response.data['summary']['rows_succeeded'], 2)
        self.assertIsNone(response.data['failures'])
        self.assertTrue(
            SchoolMembership.objects.filter(
                school=self.school,
                user__phone_number='+233244567891',
                role=User.RoleChoices.TEACHER,
                is_active=True,
            ).exists(),
        )
        self.assertEqual(
            ClassTeacher.objects.filter(
                class_level=self.class_level,
                term=self.term,
            ).count(),
            1,
        )
        self.assertEqual(
            TeachingAssignment.objects.filter(
                class_subject=self.class_subject,
                term=self.term,
            ).count(),
            1,
        )

    def test_commit_replaces_existing_assignment(self):
        teacher = create_user(
            phone_number='+233244567892',
            email='old@test.com',
            school=self.school,
            role=User.RoleChoices.TEACHER,
            is_active=True,
            first_name='Old',
            last_name='Teacher',
        )
        ClassTeacher.objects.create(
            teacher=teacher,
            class_level=self.class_level,
            term=self.term,
        )

        upload = self._csv_upload([
            {
                'first_name': 'New',
                'last_name': 'Teacher',
                'phone_number': '0244567893',
                'assignment_type': 'class_teacher',
                'class_name': 'Basic 4',
            },
        ])

        response = self.client.post(f'{self.upload_url}?dry_run=false', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['summary']['assignments_replaced'], 1)
        assignment = ClassTeacher.objects.get(class_level=self.class_level, term=self.term)
        self.assertEqual(assignment.teacher.phone_number, '+233244567893')

    @override_settings(TEACHER_BULK_IMPORT_MAX_FILE_SIZE=10)
    def test_rejects_file_larger_than_limit(self):
        upload = SimpleUploadedFile(
            'teachers.csv',
            b'x' * 20,
            content_type='text/csv',
        )

        response = self.client.post(f'{self.upload_url}?dry_run=true', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('file', str(response.data))

    def test_commit_returns_downloadable_failure_file(self):
        upload = self._csv_upload([
            {
                'first_name': 'Ama',
                'last_name': 'Boateng',
                'phone_number': '0244567891',
                'assignment_type': 'teaching',
                'class_name': 'Basic 4',
                'subject_name': 'Mathematics',
            },
            {
                'first_name': 'Bad',
                'last_name': 'Row',
                'phone_number': '0244567894',
                'assignment_type': 'teaching',
                'class_name': 'Basic 4',
                'subject_name': 'Maths',
            },
        ])

        response = self.client.post(f'{self.upload_url}?dry_run=false', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['summary']['rows_succeeded'], 1)
        self.assertEqual(response.data['summary']['rows_failed'], 1)
        self.assertIsNotNone(response.data['failures'])
        token = response.data['failures']['download_url'].rstrip('/').split('/')[-1]
        failure_path = get_failure_export_path(self.school.id, token, 'csv')
        self.assertTrue(failure_path.exists())

        download_response = self.client.get(
            reverse('school-setup-teachers-bulk-upload-failures', kwargs={'token': token}),
        )
        self.assertEqual(download_response.status_code, status.HTTP_200_OK)
        failure_content = b''.join(download_response.streaming_content).decode('utf-8-sig')
        self.assertIn('failure_reason', failure_content)

    def test_ignores_extra_columns_on_reupload(self):
        upload = self._csv_upload([
            {
                'first_name': 'Ama',
                'last_name': 'Boateng',
                'phone_number': '0244567891',
                'assignment_type': 'class_teacher',
                'class_name': 'Basic 4',
                'failure_reason': 'old error',
                'notes': 'ignore me',
            },
        ])

        response = self.client.post(f'{self.upload_url}?dry_run=false', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['summary']['rows_succeeded'], 1)

    def test_failure_xlsx_includes_dropdowns_and_reference_sheet(self):
        payload = build_teacher_failure_xlsx(
            self.school,
            [
                {
                    'first_name': 'Bad',
                    'last_name': 'Row',
                    'phone_number': '0244567894',
                    'assignment_type': 'teaching',
                    'class_name': 'Basic 4',
                    'subject_name': 'Maths',
                    'row_number': 2,
                    'failure_reason': 'Subject not found.',
                },
            ],
        )
        workbook = load_workbook(filename=io.BytesIO(payload))
        self.assertIn('Reference', workbook.sheetnames)
        import_sheet = workbook['Import']
        self.assertGreater(len(import_sheet.data_validations.dataValidation), 0)

    def test_commit_xlsx_failure_file_includes_dropdowns(self):
        upload = self._xlsx_upload([
            {
                'first_name': 'Ama',
                'last_name': 'Boateng',
                'phone_number': '0244567891',
                'assignment_type': 'teaching',
                'class_name': 'Basic 4',
                'subject_name': 'Mathematics',
            },
            {
                'first_name': 'Bad',
                'last_name': 'Row',
                'phone_number': '0244567894',
                'assignment_type': 'teaching',
                'class_name': 'Basic 4',
                'subject_name': 'Maths',
            },
        ])

        response = self.client.post(f'{self.upload_url}?dry_run=false', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['failures']['format'], 'xlsx')
        token = response.data['failures']['download_url'].rstrip('/').split('/')[-1]
        failure_path = get_failure_export_path(self.school.id, token, 'xlsx')
        workbook = load_workbook(filename=failure_path)
        self.assertIn('Reference', workbook.sheetnames)
        self.assertGreater(len(workbook['Import'].data_validations.dataValidation), 0)

    def _csv_upload(self, rows: list[dict]) -> SimpleUploadedFile:
        headers = [
            'first_name',
            'last_name',
            'phone_number',
            'email',
            'assignment_type',
            'class_name',
            'subject_name',
            'stream_name',
            'subject_group_name',
        ]
        payload = build_csv_bytes(headers=headers, rows=rows)
        return SimpleUploadedFile(
            'teachers.csv',
            payload,
            content_type='text/csv',
        )

    def _xlsx_upload(self, rows: list[dict]) -> SimpleUploadedFile:
        from schools.services.teachers_bulk_template import build_teacher_import_template

        template_bytes = build_teacher_import_template(self.school)
        workbook = load_workbook(filename=io.BytesIO(template_bytes))
        sheet = workbook['Import']
        for row in rows:
            sheet.append([
                row.get('first_name', ''),
                row.get('last_name', ''),
                row.get('phone_number', ''),
                row.get('email', ''),
                row.get('assignment_type', ''),
                row.get('class_name', ''),
                row.get('subject_name', ''),
                row.get('stream_name', ''),
                row.get('subject_group_name', ''),
            ])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            'teachers.xlsx',
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
