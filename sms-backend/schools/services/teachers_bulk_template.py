import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from schools.services.teachers_bulk_reference import build_teacher_bulk_reference_context
from shared.services.spreadsheets import (
    IMPORT_SHEET_NAME,
    LISTS_SHEET_NAME,
    REFERENCE_SHEET_NAME,
    add_list_validation,
    write_named_list,
)

IMPORT_HEADERS = [
    'first_name',
    'last_name',
    'phone_number',
    'email',
    'assignment_type',
    'class_name',
    'subject_name',
    'stream_name',
    'subject_group_name',
]

FAILURE_EXTRA_HEADERS = ['row_number', 'failure_reason']

REFERENCE_HEADERS = [
    'class_name',
    'subject_name',
    'stream_name',
    'subject_group_name',
]

ASSIGNMENT_TYPES = ['class_teacher', 'teaching']

# Dropdown columns on the Import sheet (fixed positions for import headers).
_DROPDOWN_COLUMNS = {
    'assignment_type': 'E',
    'class_name': 'F',
    'subject_name': 'G',
    'stream_name': 'H',
    'subject_group_name': 'I',
}


def build_teacher_import_template(school) -> bytes:
    context = build_teacher_bulk_reference_context(school)
    example_row = {
        'first_name': 'Ama',
        'last_name': 'Boateng',
        'phone_number': '0244567891',
        'email': 'ama@school.com',
        'assignment_type': 'class_teacher',
        'class_name': context.class_names[0] if context.class_names else 'Basic 4',
        'subject_name': '',
        'stream_name': '',
        'subject_group_name': '',
    }
    return _build_teacher_import_xlsx(
        school,
        headers=IMPORT_HEADERS,
        rows=[example_row],
    )


def build_teacher_failure_xlsx(school, failed_rows: list[dict]) -> bytes:
    headers = [*IMPORT_HEADERS, *FAILURE_EXTRA_HEADERS]
    return _build_teacher_import_xlsx(
        school,
        headers=headers,
        rows=failed_rows,
    )


def _build_teacher_import_xlsx(school, *, headers: list[str], rows: list[dict]) -> bytes:
    context = build_teacher_bulk_reference_context(school)
    workbook = Workbook()

    import_sheet = workbook.active
    import_sheet.title = IMPORT_SHEET_NAME
    import_sheet.append(headers)
    for row in rows:
        import_sheet.append([row.get(header, '') for header in headers])

    reference_sheet = workbook.create_sheet(REFERENCE_SHEET_NAME)
    reference_sheet.append(REFERENCE_HEADERS)
    for reference_row in context.reference_rows:
        reference_sheet.append([
            reference_row.class_name,
            reference_row.subject_name,
            reference_row.stream_name,
            reference_row.subject_group_name,
        ])

    lists_sheet = workbook.create_sheet(LISTS_SHEET_NAME)
    lists_sheet.sheet_state = 'hidden'

    list_ranges = {
        'assignment_type': write_named_list(
            lists_sheet,
            column=1,
            values=ASSIGNMENT_TYPES,
        ),
        'class_name': write_named_list(
            lists_sheet,
            column=2,
            values=context.class_names,
        ),
        'subject_name': write_named_list(
            lists_sheet,
            column=3,
            values=context.subject_names,
        ),
        'stream_name': write_named_list(
            lists_sheet,
            column=4,
            values=context.stream_names,
        ),
        'subject_group_name': write_named_list(
            lists_sheet,
            column=5,
            values=context.subject_group_names,
        ),
    }

    validation_max_row = max(len(rows) + 1, 1000)
    _apply_import_dropdowns(
        import_sheet,
        list_ranges=list_ranges,
        context=context,
        max_row=validation_max_row,
    )

    _autosize_sheet_columns(import_sheet, len(headers))
    _autosize_sheet_columns(reference_sheet, len(REFERENCE_HEADERS))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _apply_import_dropdowns(
    import_sheet: Worksheet,
    *,
    list_ranges: dict[str, str],
    context,
    max_row: int,
) -> None:
    add_list_validation(
        import_sheet,
        column_letter=_DROPDOWN_COLUMNS['assignment_type'],
        list_range=list_ranges['assignment_type'],
        max_row=max_row,
    )
    add_list_validation(
        import_sheet,
        column_letter=_DROPDOWN_COLUMNS['class_name'],
        list_range=list_ranges['class_name'],
        max_row=max_row,
    )
    if context.subject_names:
        add_list_validation(
            import_sheet,
            column_letter=_DROPDOWN_COLUMNS['subject_name'],
            list_range=list_ranges['subject_name'],
            max_row=max_row,
        )
    if context.stream_names:
        add_list_validation(
            import_sheet,
            column_letter=_DROPDOWN_COLUMNS['stream_name'],
            list_range=list_ranges['stream_name'],
            max_row=max_row,
        )
    if context.subject_group_names:
        add_list_validation(
            import_sheet,
            column_letter=_DROPDOWN_COLUMNS['subject_group_name'],
            list_range=list_ranges['subject_group_name'],
            max_row=max_row,
        )


def _autosize_sheet_columns(sheet: Worksheet, column_count: int) -> None:
    for index in range(1, column_count + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 22
