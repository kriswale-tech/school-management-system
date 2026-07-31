import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from shared.services.spreadsheets import (
    IMPORT_SHEET_NAME,
    LISTS_SHEET_NAME,
    REFERENCE_SHEET_NAME,
    add_list_validation,
    write_named_list,
)
from students.bulk_reference import build_student_bulk_reference_context

IMPORT_HEADERS = [
    'first_name',
    'last_name',
    'other_names',
    'gender',
    'date_of_birth',
    'admission_date',
    'is_new_student',
    'class_name',
    'guardian_name',
    'guardian_phone',
    'guardian_email',
    'guardian_relationship',
]

FAILURE_EXTRA_HEADERS = ['row_number', 'failure_reason']

REFERENCE_HEADERS = [
    'level_name',
    'class_name',
]

_DROPDOWN_COLUMNS = {
    'gender': 'D',
    'is_new_student': 'G',
    'class_name': 'H',
    'guardian_relationship': 'L',
}


def build_student_import_template(school) -> bytes:
    context = build_student_bulk_reference_context(school)
    example_class = context.class_names[0] if context.class_names else 'JHS 1'
    example_row = {
        'first_name': 'Ama',
        'last_name': 'Mensah',
        'other_names': '',
        'gender': 'female',
        'date_of_birth': '2015-03-12',
        'admission_date': '2025-09-01',
        'is_new_student': 'true',
        'class_name': example_class,
        'guardian_name': 'Akosua Mensah',
        'guardian_phone': '0244111222',
        'guardian_email': 'akosua@example.com',
        'guardian_relationship': 'mother',
    }
    return _build_student_import_xlsx(school, headers=IMPORT_HEADERS, rows=[example_row])


def build_student_failure_xlsx(school, failed_rows: list[dict]) -> bytes:
    headers = [*IMPORT_HEADERS, *FAILURE_EXTRA_HEADERS]
    return _build_student_import_xlsx(school, headers=headers, rows=failed_rows)


def _build_student_import_xlsx(school, *, headers: list[str], rows: list[dict]) -> bytes:
    context = build_student_bulk_reference_context(school)
    workbook = Workbook()

    import_sheet = workbook.active
    import_sheet.title = IMPORT_SHEET_NAME
    import_sheet.append(headers)
    for row in rows:
        import_sheet.append([row.get(header, '') for header in headers])

    reference_sheet = workbook.create_sheet(REFERENCE_SHEET_NAME)
    reference_sheet.append(REFERENCE_HEADERS)
    for reference_row in context.reference_rows:
        reference_sheet.append([
            reference_row.level_name,
            reference_row.class_name,
        ])

    lists_sheet = workbook.create_sheet(LISTS_SHEET_NAME)
    lists_sheet.sheet_state = 'hidden'

    list_ranges = {
        'gender': write_named_list(lists_sheet, column=1, values=context.gender_values),
        'is_new_student': write_named_list(
            lists_sheet,
            column=2,
            values=context.is_new_student_values,
        ),
        'class_name': write_named_list(lists_sheet, column=3, values=context.class_names),
        'guardian_relationship': write_named_list(
            lists_sheet,
            column=4,
            values=context.relationship_values,
        ),
    }

    validation_max_row = max(len(rows) + 1, 1000)
    _apply_import_dropdowns(
        import_sheet,
        list_ranges=list_ranges,
        context=context,
        max_row=validation_max_row,
    )

    _autosize_sheet_columns(import_sheet, len(headers))
    _autosize_sheet_columns(reference_sheet, len(REFERENCE_HEADERS))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _apply_import_dropdowns(
    import_sheet: Worksheet,
    *,
    list_ranges: dict[str, str],
    context,
    max_row: int,
) -> None:
    add_list_validation(
        import_sheet,
        column_letter=_DROPDOWN_COLUMNS['gender'],
        list_range=list_ranges['gender'],
        max_row=max_row,
    )
    add_list_validation(
        import_sheet,
        column_letter=_DROPDOWN_COLUMNS['is_new_student'],
        list_range=list_ranges['is_new_student'],
        max_row=max_row,
    )
    if context.class_names:
        add_list_validation(
            import_sheet,
            column_letter=_DROPDOWN_COLUMNS['class_name'],
            list_range=list_ranges['class_name'],
            max_row=max_row,
        )
    add_list_validation(
        import_sheet,
        column_letter=_DROPDOWN_COLUMNS['guardian_relationship'],
        list_range=list_ranges['guardian_relationship'],
        max_row=max_row,
    )


def _autosize_sheet_columns(sheet: Worksheet, column_count: int) -> None:
    for index in range(1, column_count + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 22
