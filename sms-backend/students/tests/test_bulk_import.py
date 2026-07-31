import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from academics.models import ClassLevel, ClassStream, Level
from accounts.tests.factories import create_user, set_client_auth_cookies, user_school
from schools.models import AcademicYear, Term
from shared.services.spreadsheets import build_csv_bytes
from students.bulk_template import IMPORT_HEADERS
from students.models import Student
from students.tests.factories import ensure_default_stream


class StudentBulkImportTests(APITestCase):
    def setUp(self):
        self.user = create_user(is_active=True)
        set_client_auth_cookies(self.client, self.user)
        self.school = user_school(self.user)
        self.school.name = 'Test Academy'
        self.school.save(update_fields=['name'])

        self.academic_year = AcademicYear.objects.create(
            school=self.school,
            academic_year='2025/2026',
            start_date='2025-09-01',
            end_date='2026-07-31',
            is_active=True,
        )
        self.term = Term.objects.create(
            school=self.school,
            academic_year=self.academic_year,
            term=Term.TermChoices.FIRST_TERM,
            start_date='2025-09-01',
            end_date='2025-12-15',
            is_active=True,
        )
        self.level = Level.objects.create(
            school=self.school,
            name='Junior High',
            is_system_generated=False,
        )
        self.class_level = ClassLevel.objects.create(
            school=self.school,
            level=self.level,
            name='JHS 1',
            is_system_generated=False,
        )
        self.default_stream = ensure_default_stream(self.class_level)
        self.named_stream = ClassStream.objects.create(
            class_level=self.class_level,
            name='A',
            is_default=False,
        )
        self.template_url = reverse('student-bulk-upload-template')
        self.upload_url = reverse('student-bulk-upload')

    def test_download_template_returns_xlsx(self):
        response = self.client.get(self.template_url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = b''.join(response.streaming_content)
        workbook = load_workbook(filename=io.BytesIO(content))
        self.assertIn('Import', workbook.sheetnames)
        self.assertIn('Reference', workbook.sheetnames)

    def test_dry_run_valid_row_without_guardian(self):
        csv_bytes = build_csv_bytes(
            headers=IMPORT_HEADERS,
            rows=[{
                'first_name': 'Kofi',
                'last_name': 'Mensah',
                'other_names': '',
                'gender': 'male',
                'date_of_birth': '2014-01-01',
                'admission_date': '2025-09-01',
                'is_new_student': '',
                'class_name': 'JHS 1 A',
                'guardian_name': '',
                'guardian_phone': '',
                'guardian_email': '',
                'guardian_relationship': '',
            }],
        )
        upload = SimpleUploadedFile('students.csv', csv_bytes, content_type='text/csv')

        response = self.client.post(
            f'{self.upload_url}?dry_run=true',
            {'file': upload},
            format='multipart',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['dry_run'])
        self.assertEqual(response.data['summary']['rows_valid'], 1)
        self.assertEqual(response.data['summary']['students_to_create'], 1)
        self.assertEqual(response.data['summary']['guardians_to_link'], 0)
        self.assertEqual(Student.objects.count(), 0)

    def test_commit_creates_student_with_optional_guardian_and_defaults(self):
        csv_bytes = build_csv_bytes(
            headers=IMPORT_HEADERS,
            rows=[
                {
                    'first_name': 'Ama',
                    'last_name': 'Boateng',
                    'other_names': '',
                    'gender': 'female',
                    'date_of_birth': '2015-03-12',
                    'admission_date': '2025-09-01',
                    'is_new_student': '',
                    'class_name': 'JHS 1 A',
                    'guardian_name': 'Akosua Boateng',
                    'guardian_phone': '0244111222',
                    'guardian_email': '',
                    'guardian_relationship': 'mother',
                },
                {
                    'first_name': 'Yaw',
                    'last_name': 'Owusu',
                    'other_names': 'Kofi',
                    'gender': 'male',
                    'date_of_birth': '2014-05-01',
                    'admission_date': '2025-09-01',
                    'is_new_student': 'true',
                    'class_name': 'JHS 1 A',
                    'guardian_name': '',
                    'guardian_phone': '',
                    'guardian_email': '',
                    'guardian_relationship': '',
                },
            ],
        )
        upload = SimpleUploadedFile('students.csv', csv_bytes, content_type='text/csv')

        response = self.client.post(
            f'{self.upload_url}?dry_run=false',
            {'file': upload},
            format='multipart',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data['dry_run'])
        self.assertEqual(response.data['summary']['rows_succeeded'], 2)
        self.assertEqual(response.data['summary']['students_created'], 2)
        self.assertEqual(response.data['summary']['guardians_linked'], 1)
        self.assertIsNone(response.data['failures'])

        ama = Student.objects.get(first_name='Ama')
        ama_enrollment = ama.enrollments.get(term=self.term)
        self.assertFalse(ama_enrollment.is_new_student)
        self.assertEqual(ama.parent_links.count(), 1)

        yaw = Student.objects.get(first_name='Yaw')
        yaw_enrollment = yaw.enrollments.get(term=self.term)
        self.assertTrue(yaw_enrollment.is_new_student)
        self.assertEqual(yaw.parent_links.count(), 0)

    def test_dry_run_reports_invalid_class(self):
        csv_bytes = build_csv_bytes(
            headers=IMPORT_HEADERS,
            rows=[{
                'first_name': 'Ama',
                'last_name': 'Mensah',
                'gender': 'female',
                'date_of_birth': '2015-03-12',
                'admission_date': '2025-09-01',
                'class_name': 'Unknown Class',
            }],
        )
        upload = SimpleUploadedFile('students.csv', csv_bytes, content_type='text/csv')

        response = self.client.post(
            f'{self.upload_url}?dry_run=true',
            {'file': upload},
            format='multipart',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['summary']['rows_with_errors'], 1)
        self.assertIn('not found', response.data['rows'][0]['messages'][0].lower())
