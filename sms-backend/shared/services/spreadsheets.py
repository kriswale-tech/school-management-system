import csv
import io
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

IMPORT_SHEET_NAME = 'Import'
REFERENCE_SHEET_NAME = 'Reference'
LISTS_SHEET_NAME = '_Lists'


def normalize_header(value: Any) -> str:
    if value is None:
        return ''
    normalized = str(value).strip().lower()
    normalized = re.sub(r'[\s\-]+', '_', normalized)
    return normalized


def normalize_cell(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def parse_xlsx_rows(
    file_bytes: bytes,
    *,
    allowed_columns: set[str],
    sheet_name: str = IMPORT_SHEET_NAME,
) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    return _rows_from_matrix(rows, allowed_columns=allowed_columns)


def parse_csv_rows(file_bytes: bytes, *, allowed_columns: set[str]) -> list[dict[str, Any]]:
    text = file_bytes.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    return _rows_from_matrix(rows, allowed_columns=allowed_columns)


def _rows_from_matrix(rows: list[tuple | list], *, allowed_columns: set[str]) -> list[dict[str, Any]]:
    if not rows:
        return []

    header_row_index = _find_header_row_index(rows, allowed_columns)
    if header_row_index is None:
        return []

    raw_headers = rows[header_row_index]
    headers = [normalize_header(value) for value in raw_headers]
    parsed_rows: list[dict[str, Any]] = []

    for offset, raw_row in enumerate(rows[header_row_index + 1 :], start=1):
        row_number = header_row_index + 1 + offset
        row_data: dict[str, Any] = {'row_number': row_number}

        has_value = False
        for index, header in enumerate(headers):
            if header not in allowed_columns:
                continue
            cell_value = normalize_cell(raw_row[index] if index < len(raw_row) else None)
            if cell_value:
                has_value = True
            row_data[header] = cell_value

        if has_value:
            parsed_rows.append(row_data)

    return parsed_rows


def _find_header_row_index(rows: list[tuple | list], allowed_columns: set[str]) -> int | None:
    for index, raw_row in enumerate(rows):
        headers = {normalize_header(value) for value in raw_row if normalize_header(value)}
        if 'phone_number' in headers and 'first_name' in headers:
            return index
        if headers.intersection(allowed_columns):
            return index
    return None


def build_xlsx_bytes(
    *,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header, '') for header in headers])

    _autosize_columns(worksheet, len(headers))
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_csv_bytes(*, headers: list[str], rows: list[dict[str, Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header, '') for header in headers})
    return buffer.getvalue().encode('utf-8-sig')


def _autosize_columns(worksheet: Worksheet, column_count: int) -> None:
    for index in range(1, column_count + 1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = 22


def add_list_validation(
    worksheet: Worksheet,
    *,
    column_letter: str,
    list_range: str,
    max_row: int = 1000,
) -> None:
    validation = DataValidation(
        type='list',
        formula1=f'={list_range}',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='Invalid value',
        error='Choose a value from the dropdown list.',
    )
    worksheet.add_data_validation(validation)
    validation.add(f'{column_letter}2:{column_letter}{max_row}')


def write_named_list(
    worksheet: Worksheet,
    *,
    column: int,
    values: list[str],
) -> str:
    letter = get_column_letter(column)
    for index, value in enumerate(values, start=1):
        worksheet.cell(row=index, column=column, value=value)
    last_row = max(len(values), 1)
    return f'{LISTS_SHEET_NAME}!${letter}$1:${letter}${last_row}'
